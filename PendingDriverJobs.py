import time
from tkinter import Tk, messagebox

from openpyxl import load_workbook


def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def show_msg(text: str, title: str = "Paving Export Count") -> None:
    root = Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    messagebox.showinfo(title, text)
    root.destroy()


def main() -> int:
    file_path = r"Z:\Chandler Projects\Paving Repair\Limited Pavement Export (Power Query).xlsx"
    sheet_name = "Paving Export"

    # 1-indexed column numbers: O=15, R=18, AA=27
    COL_O = 15   # Contractor Assigned
    COL_R = 18   # Invoice Number
    COL_AA = 27  # Print Date

    PROGRESS_EVERY = 1000  # print every N rows scanned

    start_total = time.perf_counter()
    print(f"[INFO] Opening workbook: {file_path}")

    try:
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    except FileNotFoundError:
        show_msg(f"File not found:\n{file_path}", "Error")
        return 1
    except PermissionError:
        show_msg(
            "Permission error opening the workbook.\n\n"
            "If the file is open in Excel, try closing it and run again.\n\n"
            f"Path:\n{file_path}",
            "Error",
        )
        return 1
    except Exception as e:
        show_msg(f"Failed to open workbook:\n{file_path}\n\nError: {e}", "Error")
        return 1

    open_elapsed = time.perf_counter() - start_total
    print(f"[INFO] Workbook opened in {open_elapsed:.2f}s")

    if sheet_name not in wb.sheetnames:
        show_msg(
            f"Sheet '{sheet_name}' was not found.\n\n"
            f"Available sheets:\n- " + "\n- ".join(wb.sheetnames),
            "Error",
        )
        return 1

    ws = wb[sheet_name]
    max_row = ws.max_row
    print(f"[INFO] Scanning sheet '{sheet_name}' rows: {max_row:,}")

    count = 0
    scanned = 0
    start_scan = time.perf_counter()

    # Pull only columns A..AA so we can index O/R/AA by position.
    # Start at row 2 to skip header row.
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=COL_AA, values_only=True):
        scanned += 1

        contractor_val = row[COL_O - 1]   # tuple is 0-indexed
        invoice_val = row[COL_R - 1]
        print_date_val = row[COL_AA - 1]

        contractor_ok = (
            isinstance(contractor_val, str)
            and contractor_val.strip().lower() == "driver pipeline"
        )
        invoice_blank = is_blank(invoice_val)
        print_date_present = not is_blank(print_date_val)

        if contractor_ok and invoice_blank and print_date_present:
            count += 1

        if scanned % PROGRESS_EVERY == 0:
            elapsed = time.perf_counter() - start_scan
            rate = scanned / elapsed if elapsed > 0 else 0
            # Note: scanned counts data rows (starting at Excel row 2)
            print(
                f"[PROGRESS] scanned={scanned:,}/{(max_row-1):,} | "
                f"matches={count:,} | elapsed={elapsed:.1f}s | {rate:,.0f} rows/s"
            )

    scan_elapsed = time.perf_counter() - start_scan
    total_elapsed = time.perf_counter() - start_total

    print(f"[DONE] matches={count:,} | scan={scan_elapsed:.2f}s | total={total_elapsed:.2f}s")

    show_msg(
        "Count complete.\n\n"
        "Criteria:\n"
        "- Contractor Assigned (O) = Driver Pipeline\n"
        "- Invoice Number (R) is blank\n"
        "- Print Date (AA) is present\n\n"
        f"Total matches: {count:,}\n\n"
        f"Workbook open: {open_elapsed:.2f}s\n"
        f"Scan time: {scan_elapsed:.2f}s\n"
        f"Total time: {total_elapsed:.2f}s"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
